import os

from glue.core.data_factories.helpers import has_extension
from glue.core.data_factories.pandas import panda_process
from glue.config import data_factory


__all__ = []


@data_factory(label="Excel", identifier=has_extension('xls xlsx'))
def panda_read_excel(path, sheet=None, **kwargs):
    """
    A factory for reading excel data using pandas.

    Parameters
    ----------
    path : str
        Path to the file.

    sheet : str, optional
        The sheet to read. If `None`, all sheets are read.

    **kwargs
        All other kwargs are passed to :func:`pandas.read_excel`.
    """

    try:
        import pandas as pd
    except ImportError:
        raise ImportError('Pandas is required for Excel input.')

    name = os.path.basename(path)
    if '.xls' in name:
        name = name.rsplit('.xls', 1)[0]

    if sheet is None:

        if path.endswith('xlsx'):

            try:
                from openpyxl import load_workbook
            except ImportError:
                raise ImportError('openpyxl is required for xlsx input.')

            xl_workbook = load_workbook(filename=path)

            sheet_names = xl_workbook.sheetnames

        else:

            try:
                import xlrd
            except ImportError:
                raise ImportError('xlrd is required for xls input.')

            xl_workbook = xlrd.open_workbook(path)

            sheet_names = xl_workbook.sheet_names()

    else:

        sheet_names = [sheet]

    all_data = []
    for sheet in sheet_names:
        indf = pd.read_excel(path, sheet, **kwargs)
        data = panda_process(indf)
        data.label = "{0}:{1}".format(name, sheet)
        all_data.append(data)

    return all_data
